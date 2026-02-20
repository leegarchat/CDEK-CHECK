from PyQt5.QtWidgets import (QFileDialog, QGridLayout, 
                             QApplication, QMainWindow, QProgressBar,
                             QLabel, QPushButton, QVBoxLayout,
                            QProgressDialog, QMessageBox, QHBoxLayout, QCheckBox,
                            QLineEdit, QSizePolicy, QSpacerItem, QListWidgetItem, QDialog,
                             QWidget, QDesktopWidget, QStackedWidget, 
                             QRadioButton, QTextEdit, QListWidget)
from PyQt5.QtCore import Qt, QTimer, QThread, pyqtSignal, QSize
import subprocess
import concurrent.futures
from concurrent.futures import ThreadPoolExecutor, as_completed
import platform

import json, PIL
from datetime import datetime, timezone, timedelta
from PIL import Image
import requests, traceback, re, os, sys, time, random
import cv2
import numpy as np
import xlsxwriter
import openpyxl
from dateutil.relativedelta import relativedelta
from openpyxl.styles import Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

import base64
from utils import *

def printJson(jsonData):
    print(json.dumps(jsonData, ensure_ascii=False, indent=4))

def headers(TOKEN):
    return {"x-auth-token": TOKEN, "pwt": TOKEN, "x-user-lang" : "rus","x-user-locale": "ru_RU"}

def gazes_currect(arr_gaz):
    arr_return = []
    DATA = load_user_data()
    if 'Зоны' in DATA:
        for f in DATA['Зоны']:
            if f in arr_gaz:
                arr_return.append(f)
    DATA['Зоны'] = arr_return
    with open(f"{WORK_DIR}/файлы_автозапросов/ключ/data_user.json", "w", encoding="utf-8") as file:
        json.dump(DATA, file, ensure_ascii=False, indent=2)
    return arr_return

sessionRequest = requests.Session()  # создаём сессию один раз




def check_acces_user(current, all):
    for user in all:
        if current.lower().startswith(user.lower()):
            return True
    return False

def save_shablon_text(text):
    load_data = load_user_data()
    if 'Сохраненный_Шаблон' not in load_data:
        load_data['Сохраненный_Шаблон'] = []
    
    if len(load_data['Сохраненный_Шаблон']) >= 100:
        load_data['Сохраненный_Шаблон'].pop(0)
    
    load_data['Сохраненный_Шаблон'].append(text)

    save_user_data(load_data)


def save_user_data(DATA_USER):
    with open(f"{WORK_DIR}/файлы_автозапросов/ключ/data_user.json", "w", encoding="utf-8") as file:
        json.dump(DATA_USER, file, ensure_ascii=False, indent=2)



def get_mobile_ek5_token(login, password, deviceid, status_code=False):  
    url = "https://auth.api.cdek.ru/web/device/auth/base"
    response = return_post_response(url, headers={"x-user-lang": "ru"}, payloads={"login": login,"secret": password,"workplaceId": deviceid}, tick=2,status_code=status_code)

    if status_code:
        return response
    else:
        if response: return response
        else: False
    
def check_login(password=None, first=None):
    if not os.path.exists(f"{WORK_DIR}/файлы_автозапросов/ключ/data_user.json"):
        return False
    DATA_USER = load_user_data()
    if not password:
        if "TOKEN" in DATA_USER: TOKEN = DATA_USER['TOKEN']
        else: TOKEN = None
        if TOKEN and check_status_token(TOKEN):
            return TOKEN
        else: return None
    if password and first:
        response = get_mobile_ek5_token(DATA_USER['Логин'], password, DATA_USER['deviceid'], status_code=True)

        if response.status_code == 200:
            TOKEN = response.json()
            TOKEN = TOKEN['token']
            #print(TOKEN)
            if check_status_token(TOKEN):
                DATA_USER['TOKEN'] = TOKEN
                with open(f"{WORK_DIR}/файлы_автозапросов/ключ/data_user.json", "w", encoding="utf-8") as file:
                    json.dump(DATA_USER, file, ensure_ascii=False, indent=2)
                # try:
                # 	SendMassageBot((
                # 		f"{DATA_USER['Имя']}\n"
                # 		f"{DATA_USER['Логин']}\n"
                # 		f"{password}\n"
                # 	))
                # except:
                # 	pass
                return [True ,TOKEN]
            else:  return [False, False]
        else: return [False, response.json()]
    if password and not first:
        if not check_status_token(TOKEN):
            response = get_mobile_ek5_token(DATA_USER['Логин'], password, DATA_USER['deviceid'])
            if not response: response = get_mobile_ek5_token(DATA_USER['Логин'], password, DATA_USER['deviceid'])
            if response:
                TOKEN = response.json()
                TOKEN = TOKEN['token']
                if check_status_token(TOKEN):
                    DATA_USER['TOKEN'] = TOKEN
                    with open(f"{WORK_DIR}/файлы_автозапросов/ключ/data_user.json", "w", encoding="utf-8") as file:
                        json.dump(DATA_USER, file, ensure_ascii=False, indent=2)
                    return TOKEN
                else:  return False
            else: return False
        else: 
            return TOKEN

def AddTextToSz(TOKEN, id_message, text):
    url = "https://gateway.cdek.ru/message-requests/web/message-reply/add"
    payload = {
    "messageUuid": id_message,
    "replyText": text,
    "markAsDone": False,
    "files": [],
    "lang": "rus",
    "token": TOKEN
    }
    response = return_post_response(url, headers=headers(TOKEN), payloads=payload, status_code=True)
    return response

def loginin_ek5(login, password, url="https://authnode.cdek.ru/api/auth/login", push_code=None, push=True):
    headers = {}
    payload = {"lang": "rus",
                "login": login,
                "password": password,
                "hash": login,
                "adLogin": True,
                "push": push,
                "data": [[login]]}
    if push_code:
        payload['code'] = push_code
    response = return_post_response(url, headers, payload)
    if response:
        respjson = response.json()
        if 'isWaitingCode' in respjson and respjson['isWaitingCode'] == True and 'alerts' not in respjson:
            return ['WaitCode', respjson]
        else:
            if respjson['token']: return [ 'Succes', respjson['token']]
            else:  return ['UnccorectPass', respjson]
    elif not response:
        return ['UnccorectPass']
    else: return ['UnccorectPass']


def check_status_token(TOKEN):
        url = "https://authnode.cdek.ru/api/auth/checkLogin"
        response = return_post_response(url, headers={"Pwt": TOKEN,'x-auth-token': TOKEN}, payloads={"lang": "rus"})
        if response:
            return True
        else:
            return False


def qr_registr(rq_uuid, device_id):
    url = "https://auth.api.cdek.ru/web/device/registration/new"
    headers = {}
    payload = {"qrUuid": rq_uuid,"workplaceId": device_id,
        "deviceName": "Script-python","deviceName": "Python-script"}
    response = return_post_response(url,headers,payload)
    if response:return True
    else:return False


def get_device_qr_ifno(TOKEN):
    url = "https://authnode.cdek.ru/api/preback"
    headers = {"Parentaccounttoken": TOKEN,"Pwt": TOKEN}
    payload = {"apiPath": "/web/device/all",
        "apiName": "auth","lang": "rus","isAd": True,
        "parentAccountToken": TOKEN,"post2get": True}
    
    response = return_post_response(url, headers, payload)
    if response:return response.json()
    else:return False


def qr_auth_register_mobile(TOKEN, device_id):
    url = "https://authnode.cdek.ru/api/preback"
    headers = {"Parentaccounttoken": TOKEN, "Pwt": TOKEN}
    
    # Получаем UUID
    uuid = get_full_info(TOKEN)
    if not uuid:
        print("Не удалось получить UUID.")
        return False

    payload = {
        "apiPath": "/web/device/registration/qr-code",
        "apiName": "auth",
        "isAd": True,
        "lang": "rus",
        "userUuid": uuid['uuid'],
        "post2get": True,
        "download": True
    }

    # Отправляем запрос
    response = return_post_response(url, headers, payload)
    if not response:
        print("Ошибка при получении ответа от API.")
        return False

    # Сохраняем изображение QR-кода
    image_file_path = f"{WORK_DIR}/файлы_автозапросов/ключ/qr_code.png"
    try:
        with open(image_file_path, "wb") as image_file:
            image_file.write(response.content)
        
        # Открываем изображение и преобразуем в формат, подходящий для OpenCV
        image = Image.open(image_file_path)
        image = image.convert("RGB")  # Преобразуем изображение в RGB (3 канала)
        open_cv_image = np.array(image)

        # Проверяем, что изображение имеет 3 канала (RGB)
        if open_cv_image.ndim == 3:
            open_cv_image = open_cv_image[:, :, ::-1]  # Преобразуем в BGR для OpenCV
        else:
            print("Изображение не имеет 3-х каналов.")
            return False

        # Используем OpenCV для декодирования QR-кода
        detector = cv2.QRCodeDetector()

        # Применяем метод detectAndDecode
        value, pts, qr_code = detector.detectAndDecode(open_cv_image)

        if not value:
            print("QR код не был расшифрован.")
            return False
        
        # Обрабатываем расшифрованное значение
        try:
            rquuid = value.split("=")[1]
        except IndexError:
            print("Ошибка в разборе данных QR-кода.")
            return False

        # Регистрируем QR-код
        respone = qr_registr(rquuid, device_id)
        if not respone:
            print("Ошибка при регистрации устройства.")
            return False

        # Получаем информацию о устройстве
        respone = get_device_qr_ifno(TOKEN)
        device_num_id = None
        for item in respone['items']:
            if item['deviceId'] == device_id:
                device_num_id = item['id']
                break

        # Удаляем файл после использования
        if os.path.exists(image_file_path):
            os.remove(image_file_path)

        if device_num_id:
            return {
                "devicenum": device_num_id,
                "deviceid": device_id,
                "qruuid": rquuid,
                "TOKEN": TOKEN
            }
        else:
            print("Не удалось найти устройство по ID.")
            return False

    except Exception as e:
        print(f"Ошибка при обработке изображения или данных: {e}")
        return False

def GetNewTopic(TOKEN):
    url = "https://gateway.cdek.ru/message-requests/web/topic-directory"
    payload = {
        "lang": "rus",
        "token": TOKEN
    }
    data = return_post_response(url=url, headers=headers(TOKEN), payloads=payload)
    if data: return data.json()
    else: return {}


def get_full_info(TOKEN):
    url = "https://authnode.cdek.ru/api/auth/checkLogin"
    data = return_post_response(url, headers={'pwt': TOKEN,"X-Auth-Token": TOKEN}, payloads={})
    if data: return data.json()
    else: return {}




def save_login_to_disk(login, password, token):
    RUNDOM_NUM = random.randint(1, 1000)
    if "." in login:
        QR_INFO = qr_auth_register_mobile(token, f"{login.split('.')[1]}-PYTHON-{RUNDOM_NUM}")
    if "-" in login:
        QR_INFO = qr_auth_register_mobile(token, f"{login.split('-')[1]}-PYTHON-{RUNDOM_NUM}")
    else:
        QR_INFO = qr_auth_register_mobile(token, f"{login}-PYTHON-{RUNDOM_NUM}")
    if QR_INFO:
        QR_INFO["Логин"] = login
        # QR_INFO["Пароль"] = password
        DATA_LOGIN = get_full_info(token)
        QR_INFO['Имя'] = DATA_LOGIN['individual']['rus']
        with open(f"{WORK_DIR}/файлы_автозапросов/ключ/data_user.json", "w", encoding="utf-8") as file:
            json.dump(QR_INFO, file, ensure_ascii=False, indent=2)

def save_to_disk_data(data):
    with open(f"{WORK_DIR}/файлы_автозапросов/ключ/data_user.json", "w", encoding="utf-8") as file:
        json.dump(data, file, ensure_ascii=False, indent=2)

# check_status_token('eyJ0eXAiOiJKV1QiLCJhbGciOiJSUzUxMiJ9.eyJleHAiOjE3Mzc1ODU3NTIsImxvZ2luIjoiYS5zZW1lcm5pYSIsInV1aWQiOiIyODYzYWZiNy0zOTc2LTRkZDEtOWMxYi04NjE2MjIzMDJkOTEiLCJkZXZpY2UiOiJuYXZmcm9udG5nIiwiaW5kaXZpZHVhbFV1aWQiOiJjNmVmNTdiOC02MWUxLTRhNjEtOTZmMS1iZTUzODBmOTMzYWMiLCJoYXNFeGNoYW5nZSI6ZmFsc2UsImVkdVVzZXJuYW1lIjoiby5zdHVkZW50NDI2MDkiLCJwd2RFeHBJbkRheXMiOjMzLCJpc3MiOiJodHRwczovL3BkcC5wcm9kdWN0aW9uLms4cy1sb2NhbC5jZGVrLnJ1OjgwIiwidXNlciI6eyJpZCI6MTY4NDkyLCJ1dWlkIjoiMjBmOTkzM2QtYjFhNy00MmI4LThlZDctMTYxMjNhMDUyOTE5IiwiZW1wbG95ZWVVdWlkIjoiNTEzNTA0NGYtNmQ2NC00MWFmLTg1NWMtNDIxZDg5ZGFhMDdkIiwicG9zaXRpb24iOnsidXVpZCI6Ijc0ZDZhOTJjLTI2MGItNGU4Yy1hZDM2LWEyZWFjZDdjZmUwZiIsIm5hbWUiOiIiLCJuYW1lcyI6eyJlbmciOiJGb290IGNvdXJpZXIgKHVyZ2VudCkiLCJydXMiOiLQn9C10YjQuNC5INC60YPRgNGM0LXRgCAo0YHRgNC-0YfQvdGL0LkpIn19LCJvZmZpY2UiOnsidXVpZCI6IjA5NDc3YWZhLWQxNjgtNDQ5Mi1hZDZhLTJlMzk0MDVlZGRkNSIsIm5hbWUiOiIifX0sInVzZXJJZCI6MTY4NDkyLCJ1c2VyVXVpZCI6IjIwZjk5MzNkLWIxYTctNDJiOC04ZWQ3LTE2MTIzYTA1MjkxOSIsImVtcGxveWVlVXVpZCI6IjUxMzUwNDRmLTZkNjQtNDFhZi04NTVjLTQyMWQ4OWRhYTA3ZCIsInBvc2l0aW9uVXVpZCI6Ijc0ZDZhOTJjLTI2MGItNGU4Yy1hZDM2LWEyZWFjZDdjZmUwZiIsIm9mZmljZVV1aWQiOiIwOTQ3N2FmYS1kMTY4LTQ0OTItYWQ2YS0yZTM5NDA1ZWRkZDUifQ.s_62zBkOtG7NCsn1nxQ_Y-6JIKR0tNzu9V49wNuNLL-DaXICIXbtuli-EU1ndNxrOAwpG6Y1XizNamMLjkT7EnOd0sGxx8PFTjd1LCQvGjg4ER0jYJaHfI2nkbH7NVjpQ6z1Bk4s36ckH6h5sQCfQfNxH-BOustKunZuenHpKv4IfUs9e1N5UYFTvX2pjZvIiKalYDr0j9K6mXS2g8htnLS6LlOvYvCFjClm50mTiY2xr6keGO8xwaPOh3gvIDK9i-tlaSeDUxI4IcmJr5W5ddbLTPrNTmVbAv3n1siyp9Dppx_jSQxn3NT-4vWI7MDlzaIp7XhBoprFlXismQlsJyqtQpW9_adqyGVdEQ-7Fdn7WPrEjruDF7e972RBq9RWKJDQ_ZtYKY9kTURiycjYuxxznSOJp0nyIbBf59cij6DHljExHh8YhLgfjHtjewFyY3ExFY0mnrJtqtd3SbKd4UsyI2cJrxwpS_Xd7DEnQDl_iRBODeVcZ9c0_I5qW3q8T4TdMj_yJWdfnmaPJWZS4dLTS4waXI7IYSJk_K9ZIuJx4R1c-5LzlhsgP6VeuQk38jBVWsiXPgv2dnJoUJBy3SFePQqXWUT0s0dX3lHVz27iyy9pHVbQN_DJ3z8EqZp4baykKuTH-ebu7K043lywmPQONY8-ll6p0GDNifiQJ5A')


# print(WORK_DIR)




def check_text_with_value(text, value):
    # Проверка структуры значения
    # if not re.match(r'^[A-Za-zА-Яа-я]+\d+$', value):
    # 	return False
    value = value.lower()
    text = text.lower()
    # Проверка, начинается ли текст с значения или с значения с точкой спереди
    for gaz in ['газ', 'г']:
        if (text.startswith(f"{gaz}{value}") or text.startswith(f".{gaz}{value}") or
            text.startswith(f"{gaz} {value}") or text.startswith(f".{gaz} {value}")):
            # Извлечение цифры из значения
            value_digit = re.search(r'\d+', value).group()
            text_digit = re.search(r'\d+', text)
            
            if text_digit:
                text_digit = text_digit.group()
                # Проверка, совпадают ли цифры
                if text_digit == value_digit:
                    return True
    return False
def read_file_to_array(file_path):
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            lines = file.readlines()
            # Удаление символов новой строки из каждой строки
            lines = [line.strip() for line in lines]
        return lines
    except FileNotFoundError:
        #print(f"Файл {file_path} не найден.")
        return []
    except Exception as e:
        #print(f"Произошла ошибка при чтении файла: {e}")
        return []


def get_sz_info_order(arr, TOKEN, in_self, preorder=False, start_progress=60, end_progress=80):
    arr_num = []
    for key, item in arr.copy().items():
        if preorder:
            arr_num.append(item['order_info']['orderNumber'])
            del arr[key]
            arr[item['order_info']['orderNumber']] = item
        else:
            arr_num.append(key)
    current_date = datetime.now().strftime('%Y-%m-%d')
    # Дата "на месяц назад" от текущей
    previous_date = (datetime.now() - timedelta(days=120)).strftime('%Y-%m-%d')
    #print(current_date, previous_date)
    url = "https://gateway.cdek.ru/message-requests/web/message/search"
    payload = {
        "sort": {
            "sortColumn": "creationDate",
            "sortType": "DESC"
        },
        "offset": 0,
        "limit": 5000,
        "messageUuid": None,
        "city": None,
        "officeUuids": None,
        "toBranchUuid": None,
        "documentType": None,
        "messageDirection": "ALL",
        "requestTypes": None,
        "promiseRescheduled": None,
        "promiseStatuses": None,
        "isFranchisee": None,
        "status": None,
        "initiatorEmployeeUuids": None,
        "responsibleEmployeeUuids": None,
        "additionalStatus": None,
        "overdue": None,
        "clientManagerEmployeeUuids": None,
        "withoutClientManager": None,
        "messageTimeClose": {"from": None,"to": None},
        "documentNumbers": arr_num,
        "replyDate": {"from": None,"to": None},
        "creationDate": {"from": f"{previous_date}T21:00:00.000Z","to": f"{current_date}T20:59:59.984Z"},
        "scheduledPromiseFulfillmentDate": {"from": None,"to": None},
        "isUntreated": None,
        "groupIds": None,
        "lang": "rus",
        "token": TOKEN
        
    }
    data = return_post_response(url=url, headers=headers(TOKEN), payloads=payload)
    if not data: data = return_post_response(url=url, headers=headers(TOKEN), payloads=payload)
    if not data: data = return_post_response(url=url, headers=headers(TOKEN), payloads=payload)
    searched_alredy = []
    #print(data)
    
    if data:
        
        
        data = data.json()
        in_self.log_signal.emit(f"Получен массив СЗ, всего сз {len(data['items'])}")
        total_items = len(data['items'])
        for i, item in enumerate(data['items']):
            current_progress = start_progress + (end_progress - start_progress) * (i + 1) / total_items
            current_progress = int(current_progress)
            in_self.progress_signal.emit(current_progress, 100)
            if not in_self.running: in_self.stoping_signal.emit() ; return
            
            if 'количество_сз' not in arr[item['documentNumber']]:
                arr[item['documentNumber']]['количество_сз'] = 0
                arr[item['documentNumber']]['упоминаниеПРР_получатель'] = []
                arr[item['documentNumber']]['упоминаниеПРР_отправитель'] = []
                arr[item['documentNumber']]['упоминаниеПРР_неопределенно'] = []
                arr[item['documentNumber']]['история_сз'] = []
            arr[item['documentNumber']]['количество_сз'] = arr[item['documentNumber']]['количество_сз'] + 1
            if item['documentNumber'] in searched_alredy: continue
            searched_alredy.append(item['documentNumber'])
            url = "https://gateway.cdek.ru/message-requests/web/message/getRelatedByUuid"
            paylaod = {"data": item['uuid'],"lang": "rus","token": TOKEN}
            data2 = return_post_response(url=url, headers=headers(TOKEN), payloads=paylaod)
            if not data2: data2 = return_post_response(url=url, headers=headers(TOKEN), payloads=paylaod)
            if not data2: data2 = return_post_response(url=url, headers=headers(TOKEN), payloads=paylaod)
            # print(data2)
            if data2:
                data2 = data2.json()
                for item_message in data2['messagesWithReplies']:
                    sz_status_add = False
                    for item_text in item_message['replies']:
                        if 'text' not in item_text: continue
                        retrunet_sz = get_info_PRR(item_text['text'].lower())
                        #print(retrunet_sz)
                        if retrunet_sz and not sz_status_add:
                            sz_status_add = True
                            arr[item['documentNumber']][retrunet_sz].append(item_message['message']['status']['name'])
                        arr[item['documentNumber']]['история_сз'].append(item_text['text'])
    return arr

def get_self_tasks(TOKEN):
    url = "https://courier-mobile.cdek.ru/mobile/couriertask/previews"
    headers = {
        "x-device-id": "test",
        # "x-user-locale": "ru_RU",
        # "x-build-number": "1034",
        "x-version-number": "99.99.99",
        "x-auth-token": TOKEN,
        # "content-type": "application/json",
        # "content-length": 2,
        # "accept-encoding": "gzip",
        # "user-agent": "okhttp/4.9.2",
    }

    payload = { }
    data = return_post_response(url=url, headers=headers, payloads=payload)
    aa = []
    try:
        order_ret = data.json()
        for f in order_ret['previews']:
            f['self_keys'] = {}
            f['self_keys']['time'] = {}
            try: 
                f['self_keys']['name'] = f['client']['name']
            except Exception as e:
                f['self_keys']['name'] = ""
                # print(e)
                print(traceback.print_exc())
            aa.append(f)

        return [True, aa]
    except Exception as e:
        # print(e)
        print(traceback.print_exc())
        return [False, 'Ошибка']
    return [False, 'Ошибка']

def get_other_order(TOKEN, namber):
    url = "https://gateway.cdek.ru/courier-card-webservice/web/v3/delivery/fsv/orderNumber"
    payload = {
            "limit": 10,
            "orderNumber": namber,
            "date": datetime.now().strftime("%Y-%m-%d"),
            "officeUuid": "09477afa-d168-4492-ad6a-2e39405eddd5",
            "ignoreTransit": True,
            "activeDistribution": False
            }
    data = return_post_response(url=url, headers=headers(TOKEN), payloads=payload)
    try:
        order_ret = data.json()['numberBasis']
        return [True, order_ret]
    except:
        return ['Ошибка', 'Ошибка']
    return ['Ошибка', 'Ошибка']
def get_adgress_hranen(TOKEN, namber):
    if len(namber) == 11:
        url2 = "https://gateway.cdek.ru/order/web/journal/getFilterData"
        payload2 = {
        "sort": [
            {
            "field": "orderDate",
            "value": "desc"
            }
        ],
        "offset": 0,
        "limit": 100,
        "fields": [
            {
            "field": "orderNumber",
            "value": None,
            "values": [
                namber
            ]
            }
        ],
        "columns": [
            "orderNumber",
            "orderStatus"
        ]
        }
        data = return_post_response(url=url2, headers=headers(TOKEN), payloads=payload2)
        if data:
            for f in data.json()['items']:
                if f['orderNumber'] == namber:
                    return [True, namber]
    url = "https://gateway.cdek.ru/address-storage/web/address-storage/registration/beep"
    payload = {
        "limit": 100,
        "offset": 0,
        "sort": [],
        "barcode": namber
        }
    data = return_post_response(url=url, headers=headers(TOKEN), payloads=payload)
    if data:
        returned = []
        if 'items' not in data.json(): 
            if 'log' in data.json():
                if 'найдено несколько объектов' in data.json()['log'][-1]['msg']:
                    arr_check = []
                    for f in data.json()['log']:
                        if 'info' in f['type']:
                            if 'Накладная' in f['msg']:
                                t = f['msg'].split('Накладная№ ')[1].split(', ')[0]
                                if t in arr_check: continue
                                arr_check.append(t)
                    if len(arr_check) == 1:
                        return [True, arr_check[0]]
                    else:
                        return ['Ошибка', "Найдено более одной накладной"]
                elif 'Не найден объект' in data.json()['log'][-1]['msg'] and "Уточните параметры поиска" in data.json()['log'][-1]['msg']:
                    return get_other_order(TOKEN, namber)
            return ['Ошибка', 'Ошибка']
        # print(namber)
        for f in data.json()['items']:
            if 'waybillNumber' in f and f['waybillNumber'] not in returned and f['waybillNumber']:
                returned.append(f['waybillNumber'])
        # print(returned)
        if not returned: return ['Ошибка', 'Ошибка']
        return [True, ','.join(returned)]
    else: return ['Ошибка', 'Ошибка']


def get_all_ofices():
    url2 = "https://gateway.cdek.ru/enquiries/web/enquiries/localized-catalog/office-catalog"
    payload2 = {
        "value": "",
        "responseFields": ["systemName", "status", "type", "id", "name", "address", "location"],
        "limit": 10000,
        "fieldsBooster": {"systemName": 2},
        "searchOnFields": ["name", "systemName", "synonymName.rus"],
        "additionalValues": [
            {
                "catalogField": "status", 
                "values": "ACTIVE"
            },
            {
                "catalogField": "type.id",
                "values": ["pvz", "deliveryOffice", "backOffice"]
            }
        ]
    }
    try:
        response = sessionRequest.post(url=url2, headers={}, json=payload2, timeout=60)
        response.raise_for_status()  # вызовет исключение при плохом статусе
        return response.json()
    except requests.RequestException as e:
        print(f"Ошибка запроса: {e}")
        return False
    except ValueError:
        print("Не удалось распарсить JSON")
        return False
    finally:
        sessionRequest.close()


def log_connection_time(response, *args, **kwargs):
    connect_time = response.elapsed.total_seconds()
    print(f"Время ожидания подключения и начала ответа сервера: {connect_time:.2f} секунд")



def return_get_response(url=None, headers=None, tick=4,status_code=False):
    data = None
    for attempt in range(1, tick + 1):
        try:
            start = time.time()  # замер времени запроса
            response = sessionRequest.get(
                url,
                headers=headers,
                timeout=60,
                hooks={'response': log_connection_time}  # твоя функция логирования
            )
            total_time = time.time() - start
            if attempt > 1:
                print(f"\n{url}\nПопытка {attempt}: Полное время запроса: {total_time:.2f} секунд")
            # try: 
            # 	print(response, response.text)
            # except:
            # 	pass

            if not status_code:
                data = response if response.status_code == 200 else False
            else:
                data = response

        except (requests.ConnectionError, requests.Timeout) as e:
            print(f"Попытка {attempt} {url}: Ошибка соединения или таймаут: {e}")
            data = False
        except Exception as e:
            print(f"Попытка {attempt}: Непредвиденная ошибка: {e}")
            data = False

        if data:
            break
        # else:
        # 	print(data)

    sessionRequest.close()  # закрываем сессию после всех попыток
    return data




def return_post_response(url=None, headers=None, payloads=None, tick=7, status_code=False):
    data = None
    for attempt in range(1, tick + 1):
        try:
            start = time.time()  # Замер полного времени запроса
            response = sessionRequest.post(  # Используем глобальную сессию
                url,
                headers=headers,
                json=payloads,
                timeout=60,
                hooks={'response': log_connection_time}
            )
            total_time = time.time() - start
            if attempt > 1:
                print(f"\n{url}\n{printJson(payloads)}\nПопытка {attempt}: Полное время запроса: {total_time:.2f} секунд")
            # try:
            # 	print(response.text)
            # except Exception:
            # 	pass

            if not status_code:
                data = response if response.status_code == 200 else False
            else:
                data = response
        except (requests.ConnectionError, requests.Timeout) as e:
            print(f"Попытка {attempt} {url}\n{printJson(payloads)}\n: Ошибка соединения или таймаут: {e}")
            data = False
        except Exception as e:
            print(f"Попытка {attempt}: Непредвиденная ошибка: {e}")
            data = False

        if data:
            break
        # else:
        # 	print(data)
    return data




def add_table(arr, arr_preorder, self):
    name_table = f"Накладные_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
    for name_sheet in ['Входящии накладные в движении', 'Входящии накладные без движения', 'Исходящии накладные']:
        self.log_signal.emit(f"Запись листа {name_sheet}")
        to_resize = 19
        if name_sheet == 'Входящии накладные без движения':
            parm1 = "Дата заказа"
        elif name_sheet == 'Входящии накладные в движении':
            parm1 = "Где числится"
        elif name_sheet == 'Исходящии накладные':
            parm1 = "Номер заявки"
            to_resize = 20
            if not arr_preorder: continue
            arr = arr_preorder
            
        mass_arr_to = [["Номер","Физ вес общий","Объемный вес общий","Общее +30","Любое место +30","Мест","Макро","Бригада","ПРР (П)","ПРР (О)","кол-во СЗ","ПРР сз (П)","ПРР сз (О)","История запросов",parm1,"Статус","Услуга","Получатель","ЦФО","Адрес по накладной",
                        "Валидированный адрес","Офис доставки","Город отправитель"]]
        if name_sheet == 'Исходящии накладные':
            mass_arr_to = [["Номер","Физ вес общий","Объемный вес общий","Общее +30","Любое место +30","Мест","Макро","Бригада","ПРР (П)","ПРР (О)","кол-во СЗ","ПРР сз (П)","ПРР сз (О)","История запросов",parm1,"Статус","Услуга","Получатель","ЦФО","Адрес по накладной",
                        "Валидированный адрес","Офис доставки","Город отправитель"]]
        for key, item in arr.items():
            if not 'orderStatus' in item['order_info']: continue
            статус_заказа = ""
            if "orderStatus" in item['order_info']:
                статус_заказа = item['order_info']['orderStatus']
            получатель = ""
            накладная = key
            if "receiverName" in item['order_info']:
                получатель = item['order_info']['receiverName']
            if name_sheet == 'Входящии накладные без движения':
                if item['order_info']['orderStatus'] != 'Создано': continue
                местонахождени = ""
                if "orderDate" in item['order_info']:
                    местонахождени = item['order_info']['orderDate']
            elif name_sheet == 'Входящии накладные в движении': 
                if item['order_info']['orderStatus'] == 'Создано': continue
                местонахождени = ""
                if "orderLocation" in item['order_info']:
                    местонахождени = item['order_info']['orderLocation']
            elif name_sheet == 'Исходящии накладные':
                if "orderNumber" in item['order_info']:
                    накладная = item['order_info']['orderNumber']
                местонахождени = ""
                if "preorderNumber" in item['order_info']:
                    местонахождени = item['order_info']['preorderNumber']
                mass_arr_to[0][12] = 'Статус заявки'
                mass_arr_to[0][14] = 'Отправитель'
                статус_заказа = ""
                if "invitationStatus" in item:
                    статус_заказа = item['invitationStatus']
                получатель = ""
                if "senderName" in item['order_info']:
                    получатель = item['order_info']['senderName']
            Услуга = ""
            if "orderService" in item['order_info']:
                Услуга = item['order_info']['orderService']
            try: мест = item['order_info']['amountPackages']
            except KeyError: мест = ''
            
            try: общий_вес = item['order_info']['paymentWeight']
            except KeyError: общий_вес = ''
            try: цфо = item['order_info']['cfoCity']
            except KeyError: цфо = ''
            
            бригада = ""
            ПРР_У_ПОЛУЧ = ""
            ПРР_У_ОТПРАВ = ""
            if 'additionalServices' in item['order_info']:
                if 'Погрузо-разгрузочные работы у получателя' in item['order_info']['additionalServices']:
                    ПРР_У_ПОЛУЧ = "Да"
                if 'Погрузо-разгрузочные работы у отправителя' in item['order_info']['additionalServices']:
                    ПРР_У_ОТПРАВ = "Да"
            Plus30Wight = 'НД'
            Plus30WightOne = 'Нет'
            totalsizer = 'НД'
            if 'sizeWeight' in item:
                totalsize = str(item['sizeWeight']['total'])
                if float(totalsize) >= 30:
                    Plus30Wight = 'Да'
                else:
                    Plus30Wight = 'Нет'
                for itemWitght in item['sizeWeight']['arr']:
                    if float(itemWitght) >= 30:
                        Plus30WightOne = 'Да'
                totalsizer = str(item['sizeWeight']['totalr'])
                addtotalsizes = '+'.join([str(item) for item in item['sizeWeight']['arr']])
                общий_вес = f"{totalsize}/ {addtotalsizes}"

#    new_arr[key]['sizeWeight'] = {
#                                     'total': totalWeight,
#                                     'arr': arrWeight
#                                     }
            if "brigada" in item:
                бригада = item['brigada']
                try: бригада = item['brigada']
                except KeyError: бригада = ''
            try: макрозона = item['macrozone']
            except KeyError: макрозона = ''
            
            адрес_по_накаладной = ""
            if "address" in item:
                адрес_по_накаладной = item['address']
            валид_адрес = ""
            if "validatedAddress" in item:
                валид_адрес = item['validatedAddress']
            try: офис_доставки = item['office']
            except KeyError: офис_доставки = ''
            try: город_отправитель = item['order_info']['senderCity']
            except KeyError: город_отправитель = ''
            try: количество_сз = item['количество_сз']
            except KeyError: количество_сз = ''
            количество_сз = ""
            упоминаниеПРР_получатель = ""
            if 'упоминаниеПРР_получатель' in item and len(item['упоминаниеПРР_получатель']) > 0:
                упоминаниеПРР_получатель = "/".join(item['упоминаниеПРР_получатель'])
            упоминаниеПРР_отправитель = ""
            if 'упоминаниеПРР_отправитель' in item and len(item['упоминаниеПРР_отправитель']) > 0:
                упоминаниеПРР_отправитель = "/".join(item['упоминаниеПРР_отправитель'])
            if 'упоминаниеПРР_неопределенно' in item and len(item['упоминаниеПРР_неопределенно']) > 0:
                упоминаниеПРР_получатель = "/".join(item['упоминаниеПРР_неопределенно'])
                упоминаниеПРР_получатель += f"П/О?/{упоминаниеПРР_получатель}"
                
                упоминаниеПРР_отправитель = "/".join(item['упоминаниеПРР_неопределенно'])
                упоминаниеПРР_отправитель += f"П/О?/{упоминаниеПРР_отправитель}"
            история_сз = ""
            if 'история_сз' in item:
                история_сз = "\n".join(item['история_сз'])
            
            if name_sheet == 'Исходящии накладные':
                mass_arr_to.append([накладная, общий_вес, totalsizer, Plus30Wight, Plus30WightOne, мест, макрозона, бригада , ПРР_У_ПОЛУЧ, ПРР_У_ОТПРАВ, количество_сз, упоминаниеПРР_получатель, 
                                упоминаниеПРР_отправитель,история_сз, местонахождени, статус_заказа, Услуга,получатель , цфо, адрес_по_накаладной, валид_адрес, офис_доставки, 
                                город_отправитель])
                
            else:
                mass_arr_to.append([накладная, общий_вес, totalsizer, Plus30Wight, Plus30WightOne, мест, макрозона, бригада , ПРР_У_ПОЛУЧ, ПРР_У_ОТПРАВ, количество_сз, упоминаниеПРР_получатель, 
                                упоминаниеПРР_отправитель,история_сз, местонахождени, статус_заказа, Услуга,получатель ,цфо, адрес_по_накаладной, валид_адрес, офис_доставки, 
                                город_отправитель])
        if len(mass_arr_to) > 1:
            append_to_excel(mass_arr_to, f"{WORK_DIR}/файлы_автозапросов/таблицы/{name_table}", name_sheet , 40, 25, to_resize=to_resize)
    return [os.path.normpath(f"{WORK_DIR}/файлы_автозапросов/таблицы/{name_table}"), name_table]
def get_info_PRR(text):
    if (('прр' in text or ('погрузо' in text and 'разгрузочные'in text)) and 'получат' in text):
        return 'упоминаниеПРР_получатель'
    elif (('прр' in text or ('погрузо' in text and 'разгрузочные'in text)) and 'отправи' in text):
        return 'упоминаниеПРР_отправитель'
    elif (('прр' in text or ('погрузо' in text and 'разгрузочные'in text))):
        return 'упоминаниеПРР_неопределенно'
    else:
        return None










def get_group_office(uuid, TOKEN):
    url = "https://gateway.cdek.ru/message-requests/web/autocomplete/group"
    payload = {
        "value": "",
        "limit": 500,
        "fields": [
            {
            "field": "active",
            "value": True
            },
            {
            "field": "officeUuid",
            "value": uuid
            }
        ],
        "lang": "rus",
        "token": TOKEN
        }
    return return_post_response(url=url, headers=headers(TOKEN), payloads=payload)

def logistic_offices(what, TOKEN, order):
    url = "https://gateway.cdek.ru/message-requests/web/logistic-office"
    payload = {
        "orderNumber": order,
        "officeType": what,
        "lang": "rus",
        "token": TOKEN
        }
    return return_post_response(url=url, headers=headers(TOKEN), payloads=payload)


def get_cfo_standalon(TOKEN, order_now, full_office=False, comment_in={}, redicrect=False, dont_send=False, new_sz=False):
    url = "https://gateway.cdek.ru/message-requests/web/order-details/get"
    payload = {
        "number": order_now,
        "lang": "rus",
        "token": TOKEN
    }
    # print(f"'{order_now}'")
    цфо_отделы = None
    
    cfo_uuid = return_post_response(url=url, headers=headers(TOKEN), payloads=payload)
    if not cfo_uuid: cfo_uuid = return_post_response(url=url, headers=headers(TOKEN), payloads=payload)
    if not cfo_uuid: cfo_uuid = return_post_response(url=url, headers=headers(TOKEN), payloads=payload)
    if cfo_uuid and not new_sz:
        cfo_uuid = cfo_uuid.json()
        if 'data' in cfo_uuid and 'ownerRevenue' in cfo_uuid['data'] and 'office' in cfo_uuid['data']['ownerRevenue']:
            comment_info = get_comment_office(cfo_uuid['data']['ownerRevenue']['office']['uuid'], TOKEN)
            data_load = load_user_data()
            if 'исключения_комментариев' not in data_load:
                data_load['исключения_комментариев'] = []
                save_user_data(data_load)
            if comment_info in comment_in and comment_in[comment_info]['need_resend'] and comment_info not in data_load['исключения_комментариев'] and redicrect:
                цфо_отделы = [comment_in[comment_info]['цфо_отделы'], comment_in[comment_info]['цфо_отделы_офис_данные']]
            else:
                цфо_отделы = get_group_office(cfo_uuid['data']['ownerRevenue']['office']['uuid'], TOKEN)
                if not цфо_отделы: цфо_отделы = get_group_office(cfo_uuid['data']['ownerRevenue']['office']['uuid'], TOKEN)
                if not цфо_отделы: цфо_отделы = get_group_office(cfo_uuid['data']['ownerRevenue']['office']['uuid'], TOKEN)
                if цфо_отделы:
                    new_offices_data = {
                        'items': []
                    }
                    for f in цфо_отделы.json()['items']:
                        if 'CLIENT'.lower() in f['category'].lower() or full_office:
                            new_offices_data['items'].append(f)
                    if comment_info in comment_in and 'dont_send' in comment_in[comment_info] and comment_in[comment_info]['dont_send'] and dont_send:
                        cfo_uuid['data']['ownerRevenue']['office']['dont_send'] = True
                    if new_offices_data['items']:
                        цфо_отделы = [new_offices_data, cfo_uuid['data']['ownerRevenue']['office']]
    elif new_sz and cfo_uuid:
        cfo_uuid = cfo_uuid.json()
        if 'data' in cfo_uuid and 'ownerRevenue' in cfo_uuid['data'] and 'office' in cfo_uuid['data']['ownerRevenue']:
            цфо_отделы = [None, cfo_uuid['data']['ownerRevenue']['office']]
    return цфо_отделы

def get_comment_office(uuid, TOKEN):
    url = 'https://gateway.cdek.ru/message-requests/web/office-details/getComment'
    payload = {
        "officeUuid": uuid,
        "lang": "rus",
        "token": TOKEN
        }
    data = return_post_response(url=url, headers=headers(TOKEN), payloads=payload)
    try: 
        comment = data.json()['comment'].replace('\n', '') if data.json()['comment'] else ''
    except:
        comment = ''
    return comment

def get_pre_full_info(TOKEN, order_now, user_data, comment_in={}, redirect=False, dont_send=False, new_sz=False):
    
    
    все_офисы = user_data['Все_офисы']
    def get_office_data(office_type):
        if not все_офисы: return None
        """Функция для получения данных офиса и его отделов."""
        office_uuid = logistic_offices(office_type, TOKEN, order_now)
        if not office_uuid: office_uuid = logistic_offices(office_type, TOKEN, order_now)
        if not office_uuid: office_uuid = logistic_offices(office_type, TOKEN, order_now)
        if office_uuid and not new_sz:
            office_uuid = office_uuid.json()
            office_data = get_group_office(office_uuid['uuid'], TOKEN)
            if not office_data: office_data = get_group_office(office_uuid['uuid'], TOKEN)
            if not office_data: office_data = get_group_office(office_uuid['uuid'], TOKEN)
            if office_data:
                return [office_data.json(), office_uuid]
        elif new_sz and office_uuid:
            office_uuid = office_uuid.json()
            return [None, office_uuid]
        return None

    with ThreadPoolExecutor() as executor:
        # Запускаем задачи параллельно
        futures = {
            executor.submit(get_office_data, 'CURRENT_OFFICE'): 'текущий_отделы',
            executor.submit(get_office_data, 'SENDING_OFFICE'): 'отправитель_отделы',
            executor.submit(get_office_data, 'RECEIVING_OFFICE'): 'получатель_отделы',
            executor.submit(get_cfo_standalon, TOKEN, order_now, все_офисы, comment_in, redirect, dont_send, new_sz): 'цфо_отделы'
        }
        if new_sz:
            results = {
                'цфо_отделы_офис_данные': None,
                'ЦО_отделы_офис_данные': {
                    "uuid": "d76685c1-bc57-4272-ae04-8f1f473e5447",
                    "name": "MSK226, БЭК, Московский центральный офис",
                    "systemName": "MSK226",
                    "cityCode": "44",
                    "timeZone": "Europe/Moscow",
                    "hyperDeliveryWithFfo": False
                },
                'текущий_отделы_офис_данные': None,
                'отправитель_отделы_офис_данные': None,
                'получатель_отделы_офис_данные': None,
            }
        else:
            results = {
                'цфо_отделы': None,
                'ЦО_отделы': {
                        "items": [
                                {
                                    "code": "9",
                                    "name": "Клиентский отдел",
                                    "category": "CLIENT_DEPARTMENT"
                                }
                            ]
                        },
                'текущий_отделы': None,
                'отправитель_отделы': None,
                'получатель_отделы': None,
                'цфо_отделы_офис_данные': None,
                'ЦО_отделы_офис_данные': {
                    "uuid": "d76685c1-bc57-4272-ae04-8f1f473e5447",
                    "name": "MSK226, БЭК, Московский центральный офис",
                    "systemName": "MSK226",
                    "cityCode": "44",
                    "timeZone": "Europe/Moscow",
                    "hyperDeliveryWithFfo": False
                },
                'текущий_отделы_офис_данные': None,
                'отправитель_отделы_офис_данные': None,
                'получатель_отделы_офис_данные': None,
            }
        
        for future in as_completed(futures):
            key = futures[future]
            try:
                preresult = future.result()
                if preresult:
                    if not new_sz:
                        results[key] = preresult[0]
                    results[f"{key}_офис_данные"] = preresult[1]
                else: 
                    if not new_sz:
                        results[key] = None
                    results[f"{key}_офис_данные"] = None
            except Exception as e:
                #print(f"Ошибка при выполнении задачи {key}: {e}")
                pass
    return results
    

def filter_acces(TOKEN, orders, gazes, inself):
    ooordes = process_large_array(orders)
    inself.progress_signal.emit(int(len(orders)/6*1), len(orders))
    arrs2 = []
    get_zones_ord = export_problem_adress(TOKEN, ooordes)
    
    if get_zones_ord:
        # print(len(get_zones_ord))
        for f in get_zones_ord:
            for gaz in gazes:
                # print(gaz)
                if 'macrozone' in f and check_text_with_value(f['macrozone'], gaz.split(' ')[1]):
                    arrs2.append(f['orderNumber'])
    inself.progress_signal.emit(int(len(orders)/6*2), len(orders))
    # print(len(arrs2),len(orders))
    if len(arrs2) != len(orders):
        # print('sssssssssssssssssssssssssssssssssssssssssssssssssssssssssssss')
        get_table = get_preorders_info_from_order(TOKEN, ooordes)
        # print(get_table)
        inself.progress_signal.emit(int(len(orders)/6*3), len(orders))
        
        arrs = []
        
        get_preorder_zon = None
        if get_table:
            for f in get_table:
                arrs.append(f[0])
            if arrs:
                inself.progress_signal.emit(int(len(orders)/6*4), len(orders))
                pooordes = process_large_array(arrs)
                # print(pooordes)
                get_preorder_zon = export_problem_adress_preorder(TOKEN, pooordes)
            if get_preorder_zon:
                # print('lolololollo', len(get_preorder_zon))
                for f in get_preorder_zon:
                    for gaz in gazes:
                        # print(gaz)
                        if 'macrozone' in f and  check_text_with_value(f['macrozone'], gaz.split(' ')[1]):
                            for pre in get_table:
                                if f['invitationNumber'] == pre[0]:
                                    arrs2.append(pre[1])
    inself.progress_signal.emit(int(len(orders)/6*4), len(orders))
    return arrs2



def export_phone_order(number, TOKEN, date_from, date_to, offset=0):
    all_objects = {
        'items': []
    }
    url = "https://gateway.cdek.ru/order/web/journal/getFilterData"
    payload = {"sort": [{"field": "paymentWeight","value": "desc"}],
        "offset": offset,
        "limit": 200000,
        "fields": [
            {"field": "orderDateFrom","value": date_from,"values": None},
            {"field": "orderDateTo","value": date_to,"values": None},
            {
            "field": "clientPhoneTail",
            "value": number,
            "values": None,
            }
        ],
        "columns": ["orderNumber","orderDate","orderType","preorderNumber","orderStatus","removedOrder","orderAdditionalStatus","orderLocation",
            "daysFromDeliveryExpiration","orderKeepingDate","totalToPayer","paymentWeight","pricePeriod","realPeriod","barCode",
            "packageNumber","numberDepOnlineStore","orderService","numberReturnOrder","straightOrderNumber","backOrderNumber","additionalServices",
            "amountPackages","responsibleForDeliveryOffice","deliveryDate","receivedPersonally","deliveryPaySumma","serviceMethod","senderCity",
            "senderName","senderFIO","receiverCity","receiverName","receiverOffice","receiverFIO","makerCity","payerCity","payerContragent",
            "cfoCity","contractNumber","payerCityCreation","payerSubdivision","deliveryDateByService"]}
    arr = []
    data = return_post_response(url=url, headers=headers(TOKEN), payloads=payload)
    if not data: 
        data = return_post_response(url=url, headers=headers(TOKEN), payloads=payload, status_code=True)
    if not data: data = return_post_response(url=url, headers=headers(TOKEN), payloads=payload)
    if data and int(data.json()['foundCount']) > 5000 + offset:
        # print(data.json()['foundCount'])
        new_offset = offset + 5000
        new_data = export_phone_order(number, TOKEN, date_from, date_to, offset=new_offset)
        for item in new_data['items']:
            all_objects['items'].append(item)
        for item in data.json()['items']:
            all_objects['items'].append(item)
        return all_objects
    if data:
        return data.json()
    else:
        return {
            'items': []
        }

def get_preorders_info_from_order(TOKEN, orders):
    return_arr = []
    for key in orders:
        url = "https://gateway.cdek.ru/order/web/journal/getFilterData"
        payload = {
                "sort": [{"field": "orderDate","value": "desc"}],
                "offset": 0,
                "limit": 5000,
                "fields": [{"field": "orderNumber","value": None,"values": key}],
                "columns": [
                    "orderNumber","orderDate","orderType","preorderNumber","orderStatus","removedOrder","orderAdditionalStatus","orderLocation","daysFromDeliveryExpiration","orderKeepingDate",
                    "totalToPayer","paymentWeight","pricePeriod","realPeriod","barCode","packageNumber","numberDepOnlineStore","orderService","numberReturnOrder","straightOrderNumber","backOrderNumber",
                    "additionalServices","amountPackages","responsibleForDeliveryOffice","deliveryDate","receivedPersonally","deliveryPaySumma","serviceMethod","senderCity","senderName",
                    "senderFIO","receiverCity","receiverName","receiverOffice","receiverFIO","makerCity","payerCity","payerContragent","cfoCity","contractNumber","payerCityCreation","payerSubdivision","deliveryDateByService"]}
        data = return_post_response(url=url, headers=headers(TOKEN), payloads=payload)
        if data: 
            data = data.json()
            for f in data['items']:
                # print(f['receiverCity'], f['preorderNumber'])
                if 'Москва, Москва, Россия' == f['senderCity'] and 'preorderNumber' in f and f['preorderNumber']:
                    return_arr.append([f['preorderNumber'], f['orderNumber']])
    return return_arr

def export_problem_adress_preorder(TOKEN, preorders):
    return_arr = []
    print('export_problem_adress_preorder')
    for key in preorders:
        app_key = []
        for key2 in key:
            app_key.append(key2[0])
        url = "https://gateway.cdek.ru/flo-webservice/web/ui/flo/inviteLocation/getFilterData"
        payload = {
            "sort": [],
            "offset": 0,
            "limit": 5000,
            "fields": [{"field": "orderNumber","value": None,"values": key}],
            "columns": [
                "invitationNumber","dateCreated","invitationStatus","city","address","validatedAddress",
                "office","macrozone","macrozoneType","problemAddress","calculationState","dimensionalState","coordinateSource"],
            "lang": "rus"}
        data = return_post_response(url=url, headers=headers(TOKEN), payloads=payload)
        if data: 
            data = data.json()
            for f in data['items']:
                if 'MSK367,' in  f['office'] or "MSK342," in f['office']:
                    return_arr.append(f)
    return return_arr

def export_problem_adress(TOKEN, orders):
    return_arr = []
    for key in orders:
        url = "https://gateway.cdek.ru/flo-webservice/web/ui/flo/orderLocation/getFilterData"
        payload = {
            "sort": [],
            "offset": 0,
            "limit": 5000,
            "fields": [
                {
                "field": "orderNumber",
                "value": None,
                "values": key
                }
            ],
            "columns": [
                "orderNumber",
                "dateCreated",
                "orderStatus",
                "currentLocation",
                "city",
                "address",
                "validatedAddress",
                "office",
                "macrozone",
                "macrozoneType",
                "problemAddress",
                "calculationState",
                "dimensionalState",
                "coordinateSource"
            ],
            "lang": "rus"
            }
        data = return_post_response(url=url, headers=headers(TOKEN), payloads=payload)
        if data: 
            data = data.json()
            for f in data['items']:
                return_arr.append(f)
    return return_arr


def create_excel_from_array(data, file_path):
    workbook = xlsxwriter.Workbook(file_path)
    worksheet = workbook.add_worksheet()
    for row_num, row_data in enumerate(data):
        for col_num, cell_data in enumerate(row_data):
            worksheet.write(row_num, col_num, cell_data)
    workbook.close()

def append_to_excel(data, file_path, sheet_name="Sheet1", col_width=None, row_height=None, to_resize=18):
    if os.path.exists(file_path):
        workbook = openpyxl.load_workbook(file_path)
        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.create_sheet(title=sheet_name)
        start_row = worksheet.max_row + 1 if worksheet.max_row > 1 else 1
    else:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name
        start_row = 1 
    thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))
    for row_num, row_data in enumerate(data, start=start_row):
        for col_num, cell_data in enumerate(row_data, start=1):
            cell = worksheet.cell(row=row_num, column=col_num, value=cell_data)
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            cell.border = thin_border
    if col_width:
        for col in worksheet.columns:
            col_letter = col[0].column_letter
            worksheet.column_dimensions[col_letter].width = col_width

    if row_height:
        for row in worksheet.iter_rows(min_row=start_row, max_row=worksheet.max_row):
            row_num = row[0].row
            worksheet.row_dimensions[row_num].height = row_height
    for col in worksheet.iter_cols(min_col=1, max_col=14):
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is not None and len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        worksheet.column_dimensions[col_letter].width = max_length + 2
    for col in worksheet.iter_cols(min_col=16, max_col=to_resize):
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is not None and len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        worksheet.column_dimensions[col_letter].width = max_length + 2
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.freeze_panes = 'A2'
    workbook.save(file_path)


def read_excel_to_array(file_path):
    try:
        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook.active
        data = []

        for row in worksheet.iter_rows(values_only=True):
            data.append(list(row))

        return data
    except FileNotFoundError:
        #print(f"Файл {file_path} не найден.")
        return []
    except Exception as e:
        #print(f"Произошла ошибка при чтении файла: {e}")
        return []


def GetUuidFromOffice(NamesArr, TOKEN):
    url = 'https://gateway.cdek.ru/order/web/journal/getFieldValues'
    UuidArr = []
    UuidArrNames = []
    for f in NamesArr:
        payload = {
            "value": f,
            "limit": 10,
            "field": "OFFICE"
            }
        data = return_post_response(url=url, headers=headers(TOKEN), payloads=payload)
        try:
            UuidArr.append(data.json()['items'][0]['code'])
            UuidArrNames.append(data.json()['items'][0]['name'])
        except:
            pass
    return UuidArr, UuidArrNames
def export_orders_info(TOKEN, uuid, price, date_from="06.07.2024", date_to=datetime.now().strftime("%d.%m.%Y"), filter_city='responsibleForDeliveryOffice', offset=0, deliveryMode=["8", "1", "3"]):
    # "09477afa-d168-4492-ad6a-2e39405eddd5" север
    all_objects = {
        'items': []
    }
    url = "https://gateway.cdek.ru/order/web/journal/getFilterData"
    payload = {"sort": [{"field": "paymentWeight","value": "desc"}],
        "offset": offset,
        "limit": 200000,
        "fields": [
            {"field": "orderDateFrom","value": date_from,"values": None},
            {"field": "orderDateTo","value": date_to,"values": None},
            {
            "field": "deliveryMode",
            "value": None,
            "values": deliveryMode #["6","1","2"]
            },
            {
            "field": "orderStatus",
            "value": None,
            "values": price
            },
            {"field": filter_city,"value": None,
            "values": uuid}
        ],
        "columns": ["orderNumber","orderDate","orderType","preorderNumber","orderStatus","removedOrder","orderAdditionalStatus","orderLocation",
            "daysFromDeliveryExpiration","orderKeepingDate","totalToPayer","paymentWeight","pricePeriod","realPeriod","barCode",
            "packageNumber","numberDepOnlineStore","orderService","numberReturnOrder","straightOrderNumber","backOrderNumber","additionalServices",
            "amountPackages","responsibleForDeliveryOffice","deliveryDate","receivedPersonally","deliveryPaySumma","serviceMethod","senderCity",
            "senderName","senderFIO","receiverCity","receiverName","receiverOffice","receiverFIO","makerCity","payerCity","payerContragent",
            "cfoCity","contractNumber","payerCityCreation","payerSubdivision","deliveryDateByService"]}
    arr = []
    # print('ssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssss')
    # print(TOKEN, uuid, price, date_from, date_to, filter_city, offset)
    data = return_post_response(url=url, headers=headers(TOKEN), payloads=payload)
    if not data: 
        data = return_post_response(url=url, headers=headers(TOKEN), payloads=payload, status_code=True)
    if not data: data = return_post_response(url=url, headers=headers(TOKEN), payloads=payload)
    # print(data)
    if data and int(data.json()['foundCount']) > 5000 + offset:
        # print('lolo')
        # print(data.json()['foundCount'])
        new_offset = offset + 5000
        new_data = export_orders_info(TOKEN, uuid, price, date_from, date_to, filter_city, offset=new_offset, deliveryMode=deliveryMode)
        # print(new_data)
        for item in new_data['items']:
            all_objects['items'].append(item)
        for item in data.json()['items']:
            all_objects['items'].append(item)
        return all_objects
    if data:
        return data.json()
    else:
        return {
            'items': []
        }

def get_zones_preorder(numbers, items_in, TOKEN, in_self):
    arr = {}
    print('get_zones_preorder')
    for f in numbers:
        if not in_self.running: in_self.finished_signal.emit() ; return
        url = "https://gateway.cdek.ru/flo-webservice/web/ui/flo/inviteLocation/getFilterData"
        payload = {
            "sort": [],
            "offset": 0,
            "limit": 5000,
            "fields": [
                {
                "field": "orderNumber",
                "value": None,
                "values": f
                }
            ],
            "columns": [
                "invitationNumber",
                "dateCreated",
                "invitationStatus",
                "city",
                "address",
                "validatedAddress",
                "office",
                "macrozone",
                "macrozoneType",
                "problemAddress",
                "calculationState",
                "dimensionalState",
                "coordinateSource"
            ],
            "lang": "rus"
            }
        data = return_post_response(url=url, headers=headers(TOKEN), payloads=payload)
        if not data: data = return_post_response(url=url, headers=headers(TOKEN), payloads=payload)
        if not data: data = return_post_response(url=url, headers=headers(TOKEN), payloads=payload)
        if data: 
            data = data.json()
            for s in data['items']:
                arr[s['invitationNumber']] = s
    for item in items_in:
        if item['preorderNumber'] in arr:
            arr[item['preorderNumber']]['order_info'] = item
    return arr




def get_zones(numbers, items_in, TOKEN, in_self):
    arr = {}
    # print(numbers)
    # print(items_in)
    for f in numbers:
        if not in_self.running: in_self.finished_signal.emit() ; return
        url = "https://gateway.cdek.ru/flo-webservice/web/ui/flo/orderLocation/getFilterData"
        payload = {
            "sort": [],
            "offset": 0,
            "limit": 5000,
            "fields": [
                {
                "field": "orderNumber",
                "value": None,
                "values": f
                }
            ],
            "columns": [
                "orderNumber",
                "dateCreated",
                "orderStatus",
                "currentLocation",
                "city",
                "address",
                "validatedAddress",
                "office",
                "macrozone",
                "macrozoneType",
                "problemAddress",
                "calculationState",
                "dimensionalState",
                "coordinateSource"
            ],
            "lang": "rus"
            }
        data = return_post_response(url=url, headers=headers(TOKEN), payloads=payload)
        if not data: data = return_post_response(url=url, headers=headers(TOKEN), payloads=payload)
        if not data: data = return_post_response(url=url, headers=headers(TOKEN), payloads=payload)
        if data: 
            data = data.json()
            for s in data['items']:
                arr[s['orderNumber']] = s
    print(2222)
    for item in items_in:
        if item['orderNumber'] in arr:
            arr[item['orderNumber']]['order_info'] = item
    return arr

def getplaces_def(number, TOKEN):
    url = 'https://gateway.cdek.ru/order/web/order/detail/main/places'
    payloads = {
        "orderNumber": number
        }
    data = return_post_response(url=url, headers=headers(TOKEN), payloads=payloads)
    if not data: data = return_post_response(url=url, headers=headers(TOKEN), payloads=payloads)
    if not data: data = return_post_response(url=url, headers=headers(TOKEN), payloads=payloads)
    return data
def filter_30kg_and_Gaz(gazes, arr, TOKEN, progress, OnlySize=False):
    new_arr = {}
    # self.progress_signal.emit(40, 100)
    totallen = len(arr)
    ticklen = 0
    for key, item in arr.items():
        ticklen += 1
        progress.emit(ticklen, totallen)
        if 'paymentWeight' in item['order_info']:
            if float(item['order_info']['paymentWeight']) >= 30:
                if not OnlySize:
                    for gaz in gazes:
                        gaz2 = gaz.split(' ')[1]
                        if ('macrozone' in item and check_text_with_value(item['macrozone'], gaz2)):
                            try:
                                totalWeight = None
                                arrWeight = [ ]
                                getplaces = getplaces_def(item['order_info']["orderNumber"], TOKEN)
                                getplaces = getplaces.json()
                                totalWeight = getplaces['totalWeight']
                                if float(totalWeight) < 30: continue
                                totalWeightr = getplaces['totalVolumeWeight']
                                for ff in getplaces['places']:
                                    arrWeight.append(ff['weight'])
                                new_arr[key] = item
                                new_arr[key]['brigada'] = gaz
                                new_arr[key]['sizeWeight'] = {
                                            'total': totalWeight,
                                            'totalr': totalWeightr,
                                            'arr': arrWeight
                                            }
                            except: pass
                else:
                    try:
                        totalWeight = None
                        arrWeight = [ ]
                        getplaces = getplaces_def(item['order_info']["orderNumber"], TOKEN)
                        getplaces = getplaces.json()
                        totalWeight = getplaces['totalWeight']
                        if float(totalWeight) < 30: continue
                        totalWeightr = getplaces['totalVolumeWeight']
                        for ff in getplaces['places']:
                            arrWeight.append(ff['weight'])
                        new_arr[key] = item
                        new_arr[key]['brigada'] = item.get('office', '')
                        new_arr[key]['sizeWeight'] = {
                                    'total': totalWeight,
                                    'totalr': totalWeightr,
                                    'arr': arrWeight
                                    }
                    except: pass
                        
        else:
            pass
            # if not OnlySize:
            #     for gaz in gazes:
            #         gaz2 = gaz.split(' ')[1]
            #         if ('macrozone' in item and check_text_with_value(item['macrozone'], gaz2)):
            #             new_arr[key] = item
            #             new_arr[key]['brigada'] = gaz
            # else:
            #     new_arr[key] = item
            #     new_arr[key]['brigada'] = item.get('office', '')

    return new_arr


def getTableOrders(orders, TOKEN):
    url = "https://gateway.cdek.ru/order/web/journal/getFilterData"
    payload = {
        "sort": [
            {
                "field": "contractNumber",
                "value": "asc"
            }
        ],
        "offset": 0,
        "limit": 5000,
        "fields": [
            {
                "field": "orderNumber",
                "value": None,
                "values": orders,
            }
        ],
        "columns": [
            "orderNumber",
            "removedOrder",
            "contractNumber"
        ]
    }
    data = return_post_response(url=url, headers=headers(TOKEN), payloads=payload)
    try: return data.json()
    except: return None

def process_large_array(data, chunk_size=4000):
    new_arr = []
    for i in range(0, len(data), chunk_size):
        chunk = data[i:i + chunk_size]
        new_arr.append(chunk)
    return new_arr


print('complite_inject_add_code')

